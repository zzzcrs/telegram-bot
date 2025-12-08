import os
import shutil
import tempfile

import openpyxl

from modules.marks import add_mark as db_add_mark


async def import_marks_from_excel(update, context):
    """Импорт оценок из Excel файла"""
    user_id = update.effective_user.id

    if update.message.document:
        await handle_uploaded_file(update, context, user_id)
        return
    elif context.args:
        await handle_path_file(update, context, user_id)
        return
    else:
        await show_instructions(update)


async def handle_uploaded_file(update, context, user_id):
    """Обработка загруженного файла"""
    try:
        document = update.message.document

        if not document.file_name.lower().endswith(('.xlsx', '.xls')):
            await update.message.reply_text("❌ Файл должен быть в формате Excel (.xlsx)")
            return

        await update.message.reply_text("📥 Обрабатываю файл...")

        temp_dir = tempfile.mkdtemp()
        file_path = os.path.join(temp_dir, document.file_name)

        file = await context.bot.get_file(document.file_id)
        await file.download_to_drive(file_path)

        result = await process_excel_file_ultra_simple(file_path, user_id)
        await update.message.reply_text(result)

        shutil.rmtree(temp_dir, ignore_errors=True)

    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {str(e)[:200]}")


async def handle_path_file(update, context, user_id):
    """Обработка файла по пути"""
    try:
        file_path = " ".join(context.args).strip('"\'')

        if os.path.isabs(file_path) and os.path.exists(file_path):
            result = await process_excel_file_ultra_simple(file_path, user_id)
            await update.message.reply_text(result)
        else:
            await update.message.reply_text(
                f"❌ Файл не найден или указан неполный путь.\n\n"
                f"📌 Рекомендуемый способ: отправьте файл прямо в чат 📎"
            )

    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {str(e)[:200]}")


async def process_excel_file_ultra_simple(file_path, user_id):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        added = 0
        report_lines = []
        imported_subjects = set()  
        start_row = 2 if str(sheet.cell(row=1, column=1).value).lower() in ['предмет', 'дисциплина', 'название',
                                                                            ''] else 1

        for row_idx in range(start_row, sheet.max_row + 1):
            subject_cell = sheet.cell(row=row_idx, column=1).value

            if not subject_cell:
                continue

            subject = str(subject_cell).strip()

            if not subject or subject.lower() in ['предмет', 'дисциплина', 'название']:
                continue

            subject_lower = subject.lower()

            if subject_lower in imported_subjects:
                continue

            imported_subjects.add(subject_lower)

            grades_found = []

            for col_idx in range(2, sheet.max_column + 1):
                grade_cell = sheet.cell(row=row_idx, column=col_idx).value

                if grade_cell is None:
                    continue

                try:
                    if isinstance(grade_cell, (int, float)):
                        grade = int(grade_cell)
                        if 2 <= grade <= 5:
                            grades_found.append(grade)

                    elif isinstance(grade_cell, str):
                        import re
                        numbers = re.findall(r'[2-5]', str(grade_cell))
                        for num_str in numbers:
                            grade = int(num_str)
                            if 2 <= grade <= 5:
                                grades_found.append(grade)

                except:
                    continue

            if grades_found:
                for grade in grades_found:
                    db_add_mark(user_id, subject_lower, grade)
                    added += 1

                grade_counts = {}
                for grade in grades_found:
                    grade_counts[grade] = grade_counts.get(grade, 0) + 1

                grade_strs = []
                for grade in sorted(set(grades_found)):
                    count = grade_counts[grade]
                    if count > 1:
                        grade_strs.append(f"{grade}×{count}")
                    else:
                        grade_strs.append(str(grade))

                report_lines.append(f"• {subject}: {', '.join(grade_strs)}")

        # Формируем итоговый отчет
        if added > 0:
            report = f"✅ Импортировано {added} оценок:\n\n"
            report += "\n".join(report_lines)
            return report
        else:
            return "⚠️ Не найдено оценок для импорта"

    except Exception as e:
        return f"❌ Ошибка при обработке файла: {str(e)}"


async def show_instructions(update):
    """Показать инструкции"""
    await update.message.reply_text(
        "📤 *Импорт оценок из Excel*\n\n"
        "1. 📎 *Отправьте файл* - прикрепите Excel файл к сообщению\n"
        "2. 📁 *По пути* - укажите полный путь к файлу\n\n"
        "*Требования к файлу:*\n"
        "• Первая колонка - названия предметов\n"
        "• Остальные колонки - оценки (2-5)\n\n"
        "*Пример:*\n"
        "┌──────────────┬───┬───┬───┐\n"
        "│ Математика   │ 5 │ 4 │ 5 │\n"
        "├──────────────┼───┼───┼───┤\n"
        "│ Физика       │ 4 │ 5 │   │\n"
        "└──────────────┴───┴───┴───┘",
        parse_mode='Markdown'
    )
